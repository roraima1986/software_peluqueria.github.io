from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import F
from django.shortcuts import render
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import *
from .forms import *
from .models import *
import json
from django.http import JsonResponse
import datetime
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.db.models import Q

# Lista de Categoría
class CategoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'product.view_category'
    model = Category
    template_name = 'product/category/list.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in Category.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Categoría de Productos'
        context['content_title'] = 'Lista Categoría'
        context['create_url'] = reverse_lazy('category_add')
        context['list_url'] = reverse_lazy('category_list')
        return context

# Registrar Categoria
class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'product.add_category'
    model = Category
    template_name = 'product/category/add.html'
    form_class = CategoryForm
    success_url = reverse_lazy('category_list')

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'add':
                form = self.get_form()
                if form.is_valid():
                    form.save()
                else:
                    data['error'] = form.errors
            else:
                data['error'] = 'No ha ingresado a ninguna opcion'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Categoría de Productos'
        context['content_title'] = 'Nueva Categoría'
        context['create_url'] = reverse_lazy('category_add')
        context['list_url'] = reverse_lazy('category_list')
        context['action'] = 'add'
        return context


# Editar Categoria
class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'product.change_category'
    model = Category
    template_name = 'product/category/add.html'
    form_class = CategoryForm
    success_url = reverse_lazy('category_list')

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'edit':
                form = self.get_form()
                if form.is_valid():
                    form.save()
                else:
                    data['error'] = form.errors
            else:
                data['error'] = 'No ha ingresado a ninguna opcion'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Categoría de Productos'
        context['content_title'] = 'Editar Categoría'
        context['create_url'] = reverse_lazy('category_add')
        context['list_url'] = reverse_lazy('category_list')
        context['action'] = 'edit'
        return context


# Lista de Proveedor
class ProviderListView(LoginRequiredMixin,PermissionRequiredMixin, ListView):
    permission_required = 'product.view_provider'
    model = Provider
    template_name = 'product/provider/list.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in Provider.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Proveedores'
        context['content_title'] = 'Lista Proveedor'
        context['create_url'] = reverse_lazy('provider_add')
        context['list_url'] = reverse_lazy('provider_list')
        return context


# Registrar Proveedor
class ProviderCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'product.add_provider'
    model = Provider
    template_name = 'product/provider/add.html'
    form_class = ProviderForm
    success_url = reverse_lazy('provider_list')

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'add':
                form = self.get_form()
                if form.is_valid():
                    form.save()
                else:
                    data['error'] = form.errors
            else:
                data['error'] = 'No ha ingresado a ninguna opcion'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Proveedores'
        context['content_title'] = 'Nuevo Proveedor'
        context['create_url'] = reverse_lazy('provider_add')
        context['list_url'] = reverse_lazy('provider_list')
        context['action'] = 'add'
        return context
    

# Editar Proveedor
class ProviderUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'product.change_provider'
    model = Provider
    template_name = 'product/provider/add.html'
    form_class = ProviderForm
    success_url = reverse_lazy('provider_list')

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'edit':
                form = self.get_form()
                if form.is_valid():
                    form.save()
                else:
                    data['error'] = form.errors
            else:
                data['error'] = 'No ha ingresado a ninguna opcion'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Proveedor'
        context['content_title'] = 'Editar Proveedor'
        context['create_url'] = reverse_lazy('provider_add')
        context['list_url'] = reverse_lazy('provider_list')
        context['action'] = 'edit'
        return context


# Lista de Productos
class ProductListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'product.view_product'
    model = Product
    template_name = 'product/product/list.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in Product.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Productos'
        context['content_title'] = 'Lista Productos'
        context['create_url'] = reverse_lazy('product_add')
        context['list_url'] = reverse_lazy('product_list')
        context['total_product'] = Product.objects.all().count()
        context['total_p_activos'] = Product.objects.filter(status__contains='ACTIVO').count()
        context['total_p_inhab'] = Product.objects.filter(status__contains='INHABILITADO').count()
        return context


# Registrar Productos
class ProductCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'product.add_product'
    model = Product
    template_name = 'product/product/add.html'
    form_class = ProductForm
    success_url = reverse_lazy('product_list')

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'add':
                form = self.get_form()
                if form.is_valid():
                    form.save()
                else:
                    data['error'] = form.errors
            else:
                data['error'] = 'No ha ingresado a ninguna opcion'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Productos'
        context['content_title'] = 'Nuevo Producto'
        context['create_url'] = reverse_lazy('product_add')
        context['list_url'] = reverse_lazy('product_list')
        context['action'] = 'add'
        return context


# Editar Productos
class ProductUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'product.change_product'
    model = Product
    template_name = 'product/product/add.html'
    form_class = ProductForm
    #fields = ['name', 'barcode', 'category', 'range_stock', 'price_purchase', 'price_sale', 'status', 'photo']
    success_url = reverse_lazy('product_list')

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'edit':
                form = self.get_form()
                if form.is_valid():
                    form.save()
                else:
                    data['error'] = form.errors
            else:
                data['error'] = 'No ha ingresado a ninguna opcion'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Productos'
        context['content_title'] = 'Editar Producto'
        context['create_url'] = reverse_lazy('product_add')
        context['list_url'] = reverse_lazy('product_list')
        context['action'] = 'edit'
        return context


# Reporte Excel de Productos
def report_excel_product_all(_request):
    products = Product.objects.exclude(status="INHABILITADO")

    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # cabecera
    ws['A1'] = 'NOMBRE PRODUCTO'
    ws['B1'] = 'CATEGORIA'
    ws['C1'] = 'STOCK'
    ws['D1'] = 'P/COMPRA'
    ws['E1'] = 'P/VENTA'
    ws['F1'] = 'ESTADO'

    # Iniciar en la 2da fila
    cont = 2

    # Contenido
    for product in products:
        ws.cell(row=cont, column=1).value = product.name
        ws.cell(row=cont, column=2).value = product.category.name
        ws.cell(row=cont, column=3).value = product.stock
        ws.cell(row=cont, column=4).value = product.price_purchase
        ws.cell(row=cont, column=5).value = product.price_sale
        ws.cell(row=cont, column=6).value = product.status

        cont += 1

    name_file = 'ReporteProductosExcel.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename = {}'.format(name_file)
    response['Content-Disposition'] = content
    wb.save(response)
    return response


# Reporte Excel Filtro de Productos
def report_excel_product_filter(request):
    product_name = request.GET['product_name']
    cbo_category = request.GET['cbo_category']
    cbo_stock = request.GET['cbo_stock']
    cbo_status = request.GET['cbo_status']

    # Trae todos los productos
    products = Product.objects.all()

    # Filtro de nombre de producto
    if product_name is not None:
        products = products.filter(name__icontains=product_name)

    # Filtro de categoria
    if cbo_category and cbo_category.strip():
        products = products.filter(category=cbo_category)

    # Filtro de stock de productos
    if cbo_stock == '1':
        products = products.filter(stock__gt=0)
    if cbo_stock == '2':
        products = products.filter(stock__gt=0, stock__lte=F('range_stock'))
    if cbo_stock == '3':
        products = products.filter(stock__exact=0)

    # Filtro de estado
    if cbo_status and cbo_status.strip():
        products = products.filter(status=cbo_status)

    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # cabecera
    ws['A1'] = 'NOMBRE PRODUCTO'
    ws['B1'] = 'CATEGORIA'
    ws['C1'] = 'STOCK'
    ws['D1'] = 'P/COMPRA'
    ws['E1'] = 'P/VENTA'
    ws['F1'] = 'ESTADO'

    # Iniciar en la 2da fila
    cont = 2

    # Contenido
    for product in products:
        ws.cell(row=cont, column=1).value = product.name
        ws.cell(row=cont, column=2).value = product.category.name
        ws.cell(row=cont, column=3).value = product.stock
        ws.cell(row=cont, column=4).value = product.price_purchase
        ws.cell(row=cont, column=5).value = product.price_sale
        ws.cell(row=cont, column=6).value = product.status

        cont += 1

    name_file = 'ReporteProductosExcel.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename = {}'.format(name_file)
    response['Content-Disposition'] = content
    wb.save(response)
    return response


# Valores de Categoria en el Modal
def get_categories(_request):
    categories = list(Category.objects.order_by('name').values())
    if (len(categories)>0):
        data = {
            'message':'Success',
            'categories': categories
        }
    else:
        data = {
            'message': 'No encontrado',
        }
    return JsonResponse(data)


# Lista de Compra
class BuyListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'product.view_buy'
    model = Buy
    template_name = 'product/buy/list.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in Buy.objects.all():
                    data.append(i.toJSON())
            elif action == 'search_details_prod':
                data = []
                purchases = Buy.objects.filter(id=request.POST['id'])
                for p in purchases:
                    for det in p.shopping_products:
                        data.append(det)
            elif action == 'cancel_sale':
                purchase_id = request.POST['id']
                purchase = Buy.objects.get(id=purchase_id)
                purchase.is_canceled = True
                purchase.save()
                data['success'] = 'La compra ha sido cancelada'
                # Obtener cantidad de los productos de la compra
                for prod in purchase.shopping_products:
                    id_purchase_prod = prod['id']
                    cant_purchase_prod = prod['cant']
                    # Obtener productos del inventario
                    products = Product.objects.get(id=id_purchase_prod)
                    # restar productos del inventario
                    products.stock = int(products.stock) - int(cant_purchase_prod)
                    products.save()
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Compras'
        context['content_title'] = 'Lista Compra'
        context['create_url'] = reverse_lazy('buy_add')
        context['list_url'] = reverse_lazy('buy_list')
        return context


# Registrar Compra
class BuyCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'product.add_buy'
    model = Buy
    form_class = BuyForm
    template_name = 'product/buy/add.html'
    success_url = reverse_lazy('buy_list')

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            # Buscar productos en el autocomplete
            if action == 'search_products':
                data = []
                prods = Product.objects.filter(Q(name__icontains=request.POST['term']) | Q(barcode__icontains=request.POST['term']))[0:10]
                for i in prods:
                    item = i.toJSON()
                    item['value'] = i.name
                    data.append(item)
            # Guardar productos en el modal
            elif action == 'add_product':
                prod = Product()
                prod.name = request.POST['name']
                prod.barcode = request.POST['barcode']
                prod.category = Category.objects.get(id=request.POST['category'])
                prod.range_stock = request.POST['range_stock']
                prod.status = request.POST['status']

                # En caso de que el rut del proveedor exista
                buscar = Product.objects.filter(barcode=prod.barcode).exists()
                if buscar:
                    data['error'] = 'El código de barra del producto ya existe'
                    return JsonResponse(data)

                prod.save()

            # Guardar proveedores en el modal
            elif action == 'add_provider':
                prov = Provider()
                prov.name = request.POST['name']
                prov.rut = request.POST['rut']
                prov.phone = request.POST['phone']
                prov.email = request.POST['email']
                prov.address = request.POST['address']
                prov.observation = request.POST['observation']
                # En caso de que el rut del proveedor exista
                buscar = Provider.objects.filter(rut=prov.rut).exists()
                if buscar:
                    data['error'] = 'El rut de proveedor ya existe'
                    return JsonResponse(data)
                prov.save()
                # Retornar el listado de los proveedores una vez guardados en modal
                data1 = []
                listado = Provider.objects.all()
                for i in listado:
                    prod = {
                        'id': i.id,
                        'name': i.name
                    }
                    data1.append(prod)
                return JsonResponse(data1, safe=False)

            # Guardar toda la Compra
            elif action == 'add':
                buys = json.loads(request.POST['buys'])
                provider_id = buys['provider']
                provider = Provider.objects.get(id=provider_id)
                purchase = Buy()
                purchase.provider = provider
                purchase.date_register = buys['date_register']
                purchase.n_invoice = buys['n_invoice']
                purchase.date_invoice = buys['date_invoice']
                purchase.total = int(buys['total'])
                purchase.total_prod = int(buys['total_prod'])
                purchase.shopping_products = buys['shopping_products']
                purchase.save()
                # Obtener cantidad de los productos de la compra
                for prod in purchase.shopping_products:
                    id_buy_prod = prod['id']
                    cant_buy_prod = prod['cant']
                    price_purchase_buy_prod = prod['price_purchase']
                    price_sale_buy_prod = prod['price_sale']
                    # Obtener productos del inventario
                    products = Product.objects.get(id=id_buy_prod)
                    # Sumar productos del inventario
                    products.stock = int(products.stock) + int(cant_buy_prod)
                    # Reemplazar precio de compra y precio de venta
                    products.price_purchase = price_purchase_buy_prod
                    products.price_sale = price_sale_buy_prod
                    products.save()
            else:
                data['error'] = 'No ha ingresado a ninguna opcion'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Compras'
        context['content_title'] = 'Nueva Compra'
        context['create_url'] = reverse_lazy('buy_add')
        context['list_url'] = reverse_lazy('buy_list')
        context['action'] = 'add'
        context['form_provider'] = ProviderForm
        context['form_product'] = ProductBuyForm
        return context



def report_excel_buy_all(_request):
    buys = Buy.objects.all()

    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15

    # cabecera
    ws['A1'] = 'NRO'
    ws['B1'] = 'FECHA'
    ws['C1'] = 'PROVEEDOR'
    ws['D1'] = 'RUT PROVEEDOR'
    ws['E1'] = 'N° FACTURA'
    ws['F1'] = 'FECHA FACTURA'
    ws['G1'] = 'TOTAL PRODUCTOS'
    ws['H1'] = 'TOTAL P/COMPRA'
    ws['I1'] = 'NOMBRE PRODUCTO'
    ws['J1'] = 'CANT'
    ws['K1'] = 'P/COMPRA'
    ws['L1'] = 'P/VENTA'
    ws['M1'] = 'SUBTOTAL'

    # Iniciar en la 2da fila
    cont = 2

    # Contenido
    for buy in buys:
        ws.cell(row=cont, column=1).value = buy.id
        ws.cell(row=cont, column=2).value = buy.date_register
        ws.cell(row=cont, column=3).value = buy.provider.name
        ws.cell(row=cont, column=4).value = buy.provider.rut
        ws.cell(row=cont, column=5).value = buy.n_invoice
        ws.cell(row=cont, column=6).value = buy.date_invoice
        ws.cell(row=cont, column=7).value = buy.total_prod
        ws.cell(row=cont, column=8).value = buy.total

        # Obtener valores de la lista de productos
        productos = buy.shopping_products
        # Recorrer lista de productos
        for p in productos:
            ws.cell(row=cont, column=9).value = p['name']
            ws.cell(row=cont, column=10).value = int(p['cant'])
            ws.cell(row=cont, column=11).value = int(p['price_purchase'])
            ws.cell(row=cont, column=12).value = int(p['price_sale'])
            ws.cell(row=cont, column=13).value = int(p['subtotal_prod'])
            cont += 1

        cont += 1

    name_file = 'ReporteComprasExcel.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename = {}'.format(name_file)
    response['Content-Disposition'] = content
    wb.save(response)
    return response


# Reporte Excel Filtro
def report_excel_buy_filter(request):
    radioDate = request.GET['radioDate']
    start_date = request.GET['start_date']
    finish_date = request.GET['finish_date']
    providers = request.GET['cbo_provider']

    # Traer todas las compras
    buys = Buy.objects.all()

    if radioDate == "date_invoice":
        buys = buys.filter(date_invoice__range=(start_date, finish_date))

    if radioDate == "date_system":
        buys = buys.filter(date_register__range=(start_date, finish_date))

    if providers:
        buys = buys.filter(provider=providers)

    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15

    # cabecera
    ws['A1'] = 'NRO'
    ws['B1'] = 'FECHA'
    ws['C1'] = 'PROVEEDOR'
    ws['D1'] = 'RUT PROVEEDOR'
    ws['E1'] = 'N° FACTURA'
    ws['F1'] = 'FECHA FACTURA'
    ws['G1'] = 'TOTAL PRODUCTOS'
    ws['H1'] = 'TOTAL P/COMPRA'
    ws['I1'] = 'NOMBRE PRODUCTO'
    ws['J1'] = 'CANT'
    ws['K1'] = 'P/COMPRA'
    ws['L1'] = 'P/VENTA'
    ws['M1'] = 'SUBTOTAL'

    # Iniciar en la 2da fila
    cont = 2

    # Contenido
    for buy in buys:
        ws.cell(row=cont, column=1).value = buy.id
        ws.cell(row=cont, column=2).value = buy.date_register
        ws.cell(row=cont, column=3).value = buy.provider.name
        ws.cell(row=cont, column=4).value = buy.provider.rut
        ws.cell(row=cont, column=5).value = buy.n_invoice
        ws.cell(row=cont, column=6).value = buy.date_invoice
        ws.cell(row=cont, column=7).value = buy.total_prod
        ws.cell(row=cont, column=8).value = buy.total

        # Obtener valores de la lista de productos
        productos = buy.shopping_products
        # Recorrer lista de productos
        for p in productos:
            ws.cell(row=cont, column=9).value = p['name']
            ws.cell(row=cont, column=10).value = int(p['cant'])
            ws.cell(row=cont, column=11).value = int(p['price_purchase'])
            ws.cell(row=cont, column=12).value = int(p['price_sale'])
            ws.cell(row=cont, column=13).value = int(p['subtotal_prod'])
            cont += 1

        cont += 1

    name_file = 'ReporteComprasExcel.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename = {}'.format(name_file)
    response['Content-Disposition'] = content
    wb.save(response)
    return response


# Valores de proveedores en el modal
def get_providers(request):
    providers = list(Provider.objects.order_by('name').values())
    if (len(providers)>0):
        data = {
            'message':'Success',
            'providers': providers
        }
    else:
        data = {
            'message': 'No encontrado',
        }
    return JsonResponse(data)


# Carga de productos en el menu de compras cuando se hace una buqueda
# def todos_products_compra(request):
#     print("++entro++")
#     data = list(Product.objects.all().values('id','name','barcode','price_purchase','price_sale'))
#     return  JsonResponse(data, safe=False)







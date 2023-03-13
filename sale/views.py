import datetime
import json

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.models import Group
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import CreateView, ListView
from openpyxl import Workbook
from .models import Sale
from .forms import SaleForm
from django.http import JsonResponse
from product.models import Product
from user.models import User


# Registrar venta
class SaleCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'sale.add_sale'
    model = Sale
    template_name = 'sale/add.html'
    form_class = SaleForm

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    # Para el select de los usuarios
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search_products':
                data = []
                prods = Product.objects.filter(Q(name__icontains=request.POST['term']) | Q(barcode__icontains=request.POST['term']), Q(cant__gt=0))[:10]
                for i in prods:
                    item = i.toJSON()
                    item['value'] = i.name
                    data.append(item)
            elif action == 'add':
                vents = json.loads(request.POST['vents'])
                user_id = vents['user']
                user = User.objects.get(id=user_id)
                sale = Sale()
                sale.user = user
                sale.type_sale = vents['type_sale']
                sale.observation = vents['observation']
                sale.total_sale = int(vents['total_sale'])
                sale.total_prod = int(vents['total_prod'])
                sale.output_products = vents['output_products']
                sale.save()
                # Obtener cantidad de los productos de la venta
                for prod in sale.output_products:
                    id_sale_prod = prod['id']
                    cant_sale_prod = prod['cant']
                    # Obtener productos del inventario
                    products = Product.objects.get(id=id_sale_prod)
                    # restar productos del inventario
                    products.cant = int(products.cant) - int(cant_sale_prod)
                    products.save()

            else:
                data['error'] = 'No ha ingresado a ninguna opcion'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Venta de Productos'
        context['content_title'] = 'Nueva Venta'
        context['action'] = 'add'
        context['grupo'] = Group.objects.filter(user=self.request.user).first()
        return context


# Listar las ventas
class SaleListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'sale.view_sale'
    model = Sale
    template_name = 'sale/report.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in Sale.objects.all():
                    data.append(i.toJSON())
            elif action == 'search_details_prod':
                data = []
                sales = Sale.objects.filter(id=request.POST['id'])
                for s in sales:
                    for det in s.output_products:
                        data.append(det)
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Reportes'
        context['content_title'] = 'Reporte de Venta'
        return context


# Reporte Excel de Venta (Todos las ventas)
def report_excel_sale_all(_request):
    sales = Sale.objects.exclude(is_canceled=True)

    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15

    # cabecera
    ws['A1'] = 'NRO'
    ws['B1'] = 'FECHA/HORA'
    ws['C1'] = 'PERSONAL'
    ws['D1'] = 'TIPO DE VENTA'
    ws['E1'] = 'OBSERVACIÓN'
    ws['F1'] = 'TOTAL PRODUCTOS'
    ws['G1'] = 'TOTAL P/COMPRA'
    ws['H1'] = 'CÓDIGO'
    ws['I1'] = 'NOMBRE PRODUCTO'
    ws['J1'] = 'CANT'
    ws['K1'] = 'P/VENTA'
    ws['L1'] = 'SUBTOTAL'

    # Iniciar en la 2da fila
    cont = 2

    # Contenido
    for sale in sales:
        ws.cell(row=cont, column=1).value = sale.id
        ws.cell(row=cont, column=2).value = sale.date_creation
        ws.cell(row=cont, column=3).value = sale.user.username
        ws.cell(row=cont, column=4).value = sale.type_sale
        ws.cell(row=cont, column=5).value = sale.observation
        ws.cell(row=cont, column=6).value = sale.total_prod
        ws.cell(row=cont, column=7).value = sale.total_sale

        # Obtener valores de la lista de productos
        ventas = sale.output_products
        # Recorrer lista de productos
        for v in ventas:
            ws.cell(row=cont, column=8).value = v['barcode']
            ws.cell(row=cont, column=9).value = v['name']
            ws.cell(row=cont, column=10).value = int(v['cant'])
            ws.cell(row=cont, column=11).value = int(v['price_sale'])
            ws.cell(row=cont, column=12).value = int(v['subtotal'])
            cont += 1

        cont += 1

    name_file = 'ReporteVentasExcel.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename = {}'.format(name_file)
    response['Content-Disposition'] = content
    wb.save(response)
    return response


# Reporte Excel de Venta (Filtrar ventas)
def report_excel_sale_filter(request):
    user = request.GET['cbo_user']
    chk_anulada = request.GET.get('chk_anulada', None)

    # Trae todas las ventas
    sales = Sale.objects.all()

    # Filtro por rango de fecha
    start_date_str = request.GET.get('start_date')
    finish_date_str = request.GET.get('finish_date')

    # Convierte las cadenas de texto en objetos datetime.date
    start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
    finish_date = datetime.datetime.strptime(finish_date_str, '%Y-%m-%d').date()

    # Crea los objetos de hora de inicio y fin del día
    start_time = datetime.time(00, 0, 0)
    finish_time = datetime.time(23, 59, 59, 999999)

    # Combina las fechas y horas de inicio y fin del rango
    start_datetime = datetime.datetime.combine(start_date, start_time)
    finish_datetime = datetime.datetime.combine(finish_date, finish_time)

    # Filtra las ventas por rango de fecha y hora
    sales = Sale.objects.filter(date_creation__range=(start_datetime, finish_datetime))

    # Filtro por personal
    if user:
        sales = sales.filter(user=user)

    # Filtro por ventas anuladas
    if chk_anulada:
        sales = sales.filter(is_canceled=True)
    else:
        sales = sales.filter(is_canceled=False)

    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15

    # cabecera
    ws['A1'] = 'NRO'
    ws['B1'] = 'FECHA/HORA'
    ws['C1'] = 'PERSONAL'
    ws['D1'] = 'TIPO DE VENTA'
    ws['E1'] = 'OBSERVACIÓN'
    ws['F1'] = 'TOTAL PRODUCTOS'
    ws['G1'] = 'TOTAL P/COMPRA'
    ws['H1'] = 'CÓDIGO'
    ws['I1'] = 'NOMBRE PRODUCTO'
    ws['J1'] = 'CANT'
    ws['K1'] = 'P/VENTA'
    ws['L1'] = 'SUBTOTAL'

    # Iniciar en la 2da fila
    cont = 2

    # Contenido
    for sale in sales:
        ws.cell(row=cont, column=1).value = sale.id
        ws.cell(row=cont, column=2).value = sale.date_creation
        ws.cell(row=cont, column=3).value = sale.user.username
        ws.cell(row=cont, column=4).value = sale.type_sale
        ws.cell(row=cont, column=5).value = sale.observation
        ws.cell(row=cont, column=6).value = sale.total_prod
        ws.cell(row=cont, column=7).value = sale.total_sale

        # Obtener valores de la lista de productos
        ventas = sale.output_products
        # Recorrer lista de productos
        for v in ventas:
            ws.cell(row=cont, column=8).value = v['barcode']
            ws.cell(row=cont, column=9).value = v['name']
            ws.cell(row=cont, column=10).value = int(v['cant'])
            ws.cell(row=cont, column=11).value = int(v['price_sale'])
            ws.cell(row=cont, column=12).value = int(v['subtotal'])
            cont += 1

        cont += 1

    name_file = 'ReporteVentasExcel.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename = {}'.format(name_file)
    response['Content-Disposition'] = content
    wb.save(response)
    return response


# Valores de usuarios en el modal de filtro
def get_users(_request):
    users = list(User.objects.order_by('username').values().exclude(is_superuser=True))
    if (len(users)>0):
        data = {
            'message': 'Success',
            'users': users
        }
    else:
        data = {
            'message': 'No encontrado',
        }
    return JsonResponse(data)